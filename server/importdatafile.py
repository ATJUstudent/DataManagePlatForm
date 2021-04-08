"""
Interface for import data into database by excel file <.xlsx> and csv file <.csv>
    Please look forward to more features
"""
#    Author: Wang Chuhan(wchwzhsgdx@gmail.com)
#    Time: 2021.03.14
#    for Data Manage Platform(TJU CS2018-3)
from openpyxl import load_workbook
from sqlcreator import SqlCreator
import csv
import pymysql
import datetime
import json


class FileImportTool(SqlCreator):
    """ Import data in excel file <.xlsx> and csv file <.csv>

    Interface for creating database according to excel file
    Interface for creating table in an existed database according to excel file
    Interface for inserting data into an existed table according to excel file
    Interface for inserting data into an existed table according to csv file
    Interface for rollback the new database
    Interface for rollback the new table

    """

    def deal_excel_1(self, excel, key_number=0):
        """ create database according to excel file

        Function will automatically generate database and import data according to excel data

        Parameters
        ----------
        excel: str
            path of excel file <.xlsx>
        key_number: int
            be able to specify the number of the first column as the primary key, which is the first column by default

        Notes
        -----
        Excel file name cannot be consistent with the existing database.
        You can specify which column as the primary key column, but in each workbook this number must be same.
        (The default is 0)

        """
        database = excel.split('.')[0]
        json_template = """
        {
            "database": "%s",
            "charset": "utf8mb4"
        }
        """
        json_create_database = json_template % database
        self.create_database_sql(json_create_database)
        print("Database '%s' Create Status Code:" % database, self.commit_all())

        try:
            self.deal_excel_2(excel, database, key_number)
        except ReferenceError:
            self.rollback_database_import(database)

    def deal_excel_2(self, excel, database_name, key_number=0):
        """ create table according to excel file

        Function creates the data table automatically

        Parameters
        ----------
        excel: str
            path of excel file <.xlsx>
        database_name: str
            name of an existed database
        key_number: int
            be able to specify the number of the first column as the primary key, which is the first column by default

        Notes
        -----
        You can specify which column as the primary key column, but in each workbook this number must be same.
        (The default is 0)

        """
        workbook = load_workbook(excel, data_only=True)
        sheets = workbook.sheetnames

        for sheet in sheets:
            field_template = {}
            json_template = {sheet: field_template}
            work_sheet = workbook[sheet]
            value_row = []
            for row in work_sheet.rows:
                value = [cell.value for cell in row]
                value_row.append(value)
            num = 0
            for v in value_row[1]:
                if num == key_number:
                    e = 'PRI'
                else:
                    e = ''

                if type(v) == int:
                    field_template[num] = {'Field': value_row[0][num], 'Type': 'INT', 'Key': e}
                elif type(v) == float:
                    field_template[num] = {'Field': value_row[0][num], 'Type': 'FLOAT', 'Key': e}
                elif type(v) == str:
                    field_template[num] = {'Field': value_row[0][num], 'Type': 'VARCHAR(255)', 'Key': e}
                elif type(v) == datetime.datetime:
                    field_template[num] = {'Field': value_row[0][num], 'Type': 'DATETIME', 'Key': e}
                elif type(v) == datetime.date:
                    field_template[num] = {'Field': value_row[0][num], 'Type': 'DATE', 'Key': e}
                elif type(v) == datetime.time:
                    field_template[num] = {'Field': value_row[0][num], 'Type': 'TIME', 'Key': e}
                else:
                    raise TypeError('不支持的数据类型！')
                num = num + 1
            json_create_table = json.dumps(json_template)
            self.create_table_sql(json_create_table, database_name)
            print("Table '%s' Create Status Code:" % sheet, self.commit_all())

            num = 0
            json_template = {}
            for values in value_row[1:]:
                value = [v for v in values]
                value_num = len(value)
                value_d = {}
                for i in range(value_num):
                    if value[i] is None:
                        self.rollback_table_import(database_name, sheet)
                        raise ReferenceError('表中数据存在空存档，请修改后重试')

                    if type(value[i]) == datetime.datetime or type(value[i]) == datetime.date \
                            or type(value[i]) == datetime.time:
                        value_d[value_row[0][i]] = str(value[i]).split('.')[0]
                    else:
                        value_d[value_row[0][i]] = value[i]
                json_template['%d' % num] = value_d
                num = num + 1
            json_insert_table = json.dumps(json_template, ensure_ascii=False)
            self.create_object_sql(json_insert_table, database_name, sheet)
            print("Insert Into Table '%s' Status Code:" % sheet, self.commit_all())

    def deal_excel_3(self, excel, database_name, table_name, sheet_seq=0):
        """ insert data into table according to excel file <.xlsx>

        Function automatically matches column and database properties by name

        Parameters
        ----------
        excel: str
            path of excel file <.xlsx>
        database_name: str
            name of an existed database
        table_name: str
            name of an existed table
        sheet_seq: int
            the serial number of the imported workbook

        Notes
        -----
        if not specify the serial number of imported workbook, it will be defaulted by 0 (the first work sheet)

        """
        workbook = load_workbook(excel, data_only=True)
        sheets = workbook.sheetnames
        sheet = workbook[sheets[sheet_seq]]
        value_row = []
        for row in sheet.rows:
            value = [cell.value for cell in row]
            value_row.append(value)
        self.insert_value_row(value_row, database_name, table_name)

    def deal_csv(self, csv_file, database_name, table_name):
        """ insert data into table according to csv file <.csv>

        Function automatically matches column and database properties by name

        Parameters
        ----------
        csv_file: str
            path of csv file <.csv>
        database_name: str
            name of an existed database
        table_name: str
            name of an existed table

        """
        with open(csv_file) as f_csv:
            file = csv.reader(f_csv)
            value_row = []
            for row in file:
                value = [v.replace(' ', '') for v in row]
                value_row.append(value)
        f_csv.close()
        self.insert_value_row(value_row, database_name, table_name)

    def insert_value_row(self, value_row, database_name, table_name):
        """ inserts the input data into the specified table

        Parameters
        ----------
        value_row: list
            the input data
        database_name: str
            name of an existed database
        table_name: str
            name of an existed table

        Notes
        -----
        the value of value_row[0] must be the name of column

        """
        input_name = []
        for value in value_row[0]:
            input_name.append(value)
        accept_name = []
        for key in self.table_columns(database_name, table_name).fetchall():
            accept_name.append(key['Field'])

        try:
            auto_map = my_match_list(input_name, accept_name)  # <input -> accept>
        except ValueError:
            raise ReferenceError('两张表所含数据名称不一一对应！')
        except KeyError:
            raise ReferenceError('两张表所含数据名称不一一对应！')
        except ReferenceError:
            raise Warning('输入数据不能对所有数据库属性赋值，可能会产生意想不到的错误！')

        num = 0
        json_template = {}
        for values in value_row[1:]:
            value = [v for v in values]
            value_num = len(value)
            value_d = {}
            for i in range(value_num):
                if value[i] is None:
                    raise ReferenceError('表中数据存在空存档，请修改后重试')

                if type(value[i]) == datetime.datetime or type(value[i]) == datetime.date \
                        or type(value[i]) == datetime.time:
                    value_d[value_row[0][auto_map[i]]] = str(value[i]).split('.')[0]
                else:
                    value_d[value_row[0][auto_map[i]]] = value[i]
            json_template['%d' % num] = value_d
            num = num + 1
        json_insert_table = json.dumps(json_template, ensure_ascii=False)
        self.create_object_sql(json_insert_table, database_name, table_name)
        print(json_insert_table)
        print("Insert Into Table '%s' Status Code:" % table_name, self.commit_all())

    def rollback_database_import(self, database_name):
        try:
            self.commit_sql('DROP DATABASE %s;' % database_name)
        except pymysql.err.Error:
            print('表中数据存在空存档，请修改后重试')
            raise ReferenceError('Rollback Import Error: Cannot drop database.')

    def rollback_table_import(self, database_name, table_name):
        try:
            self.commit_sql('DROP TABLE %s.%s;') % (database_name, table_name)
        except pymysql.err.Error:
            print('表中数据存在空存档，请修改后重试')
            raise ReferenceError('Rollback Import Error: Cannot drop table.')


def my_match_list(list1, list2):
    list2_stack = {}
    i = 0
    for el in list2:
        try:
            list2_stack[el].append(list2.index(el, i, len(list2)))
            i = i + 1
        except KeyError:
            list2_stack[el] = []
            list2_stack[el].append(list2.index(el, i, len(list2)))
            i = i + 1

    result = {}
    i = 0
    for el in list1:
        try:
            result[i] = list2_stack[el].pop(0)
            i = i + 1
        except ValueError:
            raise ValueError('Value Error: The number of elements in LIST1 and List2 is not equal!')
        except KeyError:
            raise ValueError('Key Error: There is no such element in List2!')
    for key in list2_stack:
        if list2_stack[key]:
            raise ReferenceError('Reference Error: The two lists are not equal in length!')

    return result


# # 测试用代码，取消注释使用
# if __name__ == '__main__':
#     dict1 = {
#         "ip": "127.0.0.1",
#         "port": 3306,
#         "database": "test",
#         "username": "root",
#         "password": "123456"
#     }
#     FileImportTool.init_config(dict1)
#     it = FileImportTool()
#     it.connect_db()
#     # print(it.deal_excel_1(u'TotalData.xlsx'))
#     it.rollback_database_import('totaldata')
#     # print(it.deal_excel_2(u'TotalData.xlsx', 'totaldata'))
#     # it.deal_excel_3(u'TotalData.xlsx', 'totaldata', 'sheet1', sheet_seq=2)
#     # it.deal_csv('TotalData.csv', 'totaldata', 'sheet3')
